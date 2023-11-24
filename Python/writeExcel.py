#! python3 update Spreadsheet

#1. Loops over all the rows
#2. If the row is for garlic, celery, or lemons, changes the price

import openpyxl

#1. Open the spreadsheet file.
print('Opening workbook')
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# initialise data
# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
 'Celery': 1.19,
 'Lemon': 1.27}

#2. For each row, check whether the value in column A is Celery, Garlic,
#or Lemon.
# Loop through the rows and update the prices
for rowNum in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row=rowNum, column=1).value
    print('Got ' + produceName)
    #3. If it is, update the price in column B.
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

#4. Save the spreadsheet to a new file (so that you don’t lose the old spreadsheet, just in case).
wb.save('updateProduceSales.xlsx')

